import openpyxl
from datetime import datetime

today = datetime.today()
today_time = today.strftime('%Y/%m/%d')
today_time2 = today.strftime('%Y%m%d')


wb = openpyxl.Workbook()
ws = wb.active

# セルに値を入力
data = [
    ["請求書"],
    ["株式会社ABC","","","","No.","0001"],
    ["〒101-0022 東京都千代田区神田練塀町300","","","","日付",today_time],
    ["TEL:03-1234-5678 FAX:03-1234-5678"],
    ["担当者名:鈴木一郎 様"],
    ["商品名","数量","単価","金額"],
    ["商品A",2,10000,20000],
    ["商品B",1,15000,15000],
    ["合計"],
    ["消費税"],
    ["税込合計"],
]

for row in data:
  ws.append(row)


# 指定した行に新しい行を挿入
ws.insert_rows(1, 1)
ws.insert_rows(3, 1)
ws.insert_rows(8, 2)
ws.insert_rows(13, 2)
ws.insert_cols(1, 1)


# 数式の挿入
ws["E13"] = "=E11+E12"
ws["E15"] = "=E13"
ws["E16"] = "=E15*0.1"
ws["E17"] = "=E15+E16"

# セルの幅調整
seru = ["B","C","D","E"]

for i in seru:
  ws.column_dimensions[i].width = 11

# ファイル保存
wb.save(f"請求書_{today_time2}.xlsx")
