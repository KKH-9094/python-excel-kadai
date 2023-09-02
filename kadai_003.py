import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

wb1 = pd.read_excel('2022_年間売上表.xlsx')
df1 = pd.DataFrame(wb1)
# print(df1)

wb2 = pd.read_excel('2023_年間売上表.xlsx')
df2 = pd.DataFrame(wb2)
# print(df2)

df3 = pd.concat([df1,df2])

re = df3.groupby(["商品","売上年"])["金額（千円）"].sum().reset_index()
# print(re)

re.to_excel('売上集計表.xlsx',sheet_name='Sheet1', index=False)

wb3 = openpyxl.load_workbook('売上集計表.xlsx')

# ヘッダー部分（1行目）を薄いグレー（#F2F2F2）に設定し

ws = wb3.active

co = ["A","B","C"]

for i in co:
  ws[f'{i}1'].fill = PatternFill(patternType='solid', fgColor='F2F2F2')

seru = ["A","B","C","D","E"]

for i in seru:
  ws.column_dimensions[i].width = 11

wb3.save('売上集計表.xlsx')
